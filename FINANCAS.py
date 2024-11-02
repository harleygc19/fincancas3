import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
import datetime

st.set_page_config(page_title="Finanças HD", layout='wide', page_icon="logo2.png")

# Carregar o arquivo Contas.xlsx
df = pd.read_excel('Contas.xlsx')
#df['Data do pagamento'] = pd.to_datetime(df['Data do pagamento']).dt.strftime('%d/%m/%y')


# Função para verificar credenciais
def verificar_credenciais(username, senha):
    try:
        df = pd.read_excel('senhas.xlsx')
    except FileNotFoundError:
        st.error('Arquivo senhas não encontrado')        
        return False
    except Exception as e:
        st.error(f'Erro ao ler o arquivo: {e}')
        return False

    if any((df["username"] == username) & (df["senha"] == senha)):
        return True
    return False

# Função para formatar datas
def formatar_datas(df):
    colunas_de_data = ['Data_vencimento', 'Data_do_pagamento'] 
    for coluna in colunas_de_data:
        if coluna in df.columns:
            df[coluna] = pd.to_datetime(df[coluna], errors='coerce').dt.strftime('%d/%m/%Y')
    return df

# Ajustando o DataFrame inicial para exibir datas no formato desejado
df = formatar_datas(df)

# Página de login
def login_page():
    col1, col2, col3 = st.columns((0.5, 1, 0.5))
    with col2:
        st.image('logo1.png', use_column_width=True)
        st.sidebar.image('logo3.png', use_column_width=True)
    
    username = st.sidebar.text_input('login')
    senha = st.sidebar.text_input('senha', type='password')

    if st.sidebar.button('login'):
        if verificar_credenciais(username, senha):
            st.session_state['logged_in'] = True
            st.session_state['username'] = username
            st.rerun()
        else: 
            st.error('usuário ou senha incorretos, tente novamente')
    st.sidebar.text('Controle financeiro HD') 

# Página Principal
def pag_inicial():
    st.subheader('Controle Financeiro HD', divider='gray')
    pag1, pag2 = st.tabs(['Cadastro de Gastos', 'Cadastro de Recebimentos'])

    with pag1:
        col1, col2 = st.columns(2)
        st.write(df)
        formulario_c_gastos1()
        nova_categoria()
        Nova_conta()

    with pag2:
        df_recebimentos = pd.read_excel('Recebimentos.xlsx')
        st.dataframe(formatar_datas(df_recebimentos), hide_index=True)
        recebimento()

# Função para carregar e formatar o DataFrame
def formulario():
    try:
        df = formatar_datas(df)
        df = pd.read_excel('Contas.xlsx')
        
        #df = formatar_datas(df)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Categoria', 'Conta', 'Forma_de_pagamento', 'Data_do_pagamento', 'Status_do_pagamento', 'Valor_pago', 'OBS'])
    
    return df

# Página Principal
def pag_inicial():
    st.subheader('Controle Financeiro HD',divider='gray')
    df['Data do Pagamento'] = pd.to_datetime(df['Data do Pagamento'])
    pag1, pag2 = st.tabs(['Cadastro de Gastos','Cadastro de Recebimentos'])

    with pag1:
        col1, col2,col3 = st.columns(3)

        with col2:
            # Inputs de data
            pos1,pos2 = st.columns(2)
            with pos1:
                start_date = st.date_input("Data de Início", datetime.date(2024, 11, 1), key='DataUniIni')
            with pos2:
                end_date = st.date_input("Data de Fim", datetime.date(2024, 12, 31), key='DataOperFim')

            # Conversão para datetime (apenas uma vez)
            start_date = pd.to_datetime(start_date)
            end_date = pd.to_datetime(end_date)

            # Verificação da validade do intervalo de datas
            if start_date > end_date:
                st.error("Data de Início não pode ser maior que a Data de Fim.")
            else:
                # Formatação da coluna 'Data do pagamento' como datetime para realizar o filtro
                df['Data do Pagamento'] = pd.to_datetime(df['Data do Pagamento'], errors='coerce')
                
                # Filtrando o DataFrame pelo intervalo de datas selecionado
                df_filtered = df[(df['Data do Pagamento'] >= start_date) & (df['Data do Pagamento'] <= end_date)]
                
                # Exibindo o DataFrame filtrado
                st.write(df_filtered,hide_index=True)

            formulario_c_gastos1()
            nova_categoria()
            Nova_conta()

    with pag2:
        df_recebimentos = pd.read_excel('Recebimentos.xlsx')
        st.dataframe(df_recebimentos,hide_index=True)
        recebimento()
        pass 

#Criando o Cadastro de gastos
df = pd.read_excel('Contas.xlsx')

def formulario_c_gastos1():
    forma_de_pagamento_lista = 'Debito Harley','Debito Daiana','Credito Bradesco','Credito Nubank','Crédito Outros'
    @st.dialog('Cadastrar nova conta',width='big')
    def formulario_c_gastos():    
        
        lista_status = 'Pago','Pendente'
        
        dfcat = pd.read_excel('nome_categorias.xlsx')
        dfcat = dfcat[list]
        
        dfconta = pd.read_excel('nome_contas.xlsx')
        dfconta = dfconta[list]
        
        lista_categoria = dfcat
        lista_contas = dfconta

        if ' Categoria' not in st.session_state:
            st.session_state['Categoria'] = ''
        if 'Conta' not in st.session_state:
            st.session_state['Conta'] = ''
        if ' Forma de Pagamento' not in st.session_state:
            st.session_state['Forma_de_pagamento'] = ''
        if 'Data do pagamento' not in st.session_state:
            st.session_state['Data_do_pagamento'] = ''
        if 'Status do pagamento' not in st.session_state:
            st.session_state['Status_do_pagamento'] = ''
        if 'Valor_pago' not in st.session_state:
            st.session_state['Valor_pago'] = ''
        if 'OBS' not in st.session_state:
            st.session_state['OBS'] = ''
    
        st.session_state['Categoria'] = st.selectbox('Categoria',lista_categoria)
        st.session_state['Conta'] = st.selectbox('Conta', lista_contas)
        st.session_state['Forma_de_pagamento']=st.selectbox('Forma de pagamento',forma_de_pagamento_lista)
        st.session_state['Data_do_pagamento']=st.date_input('Data do pagamento')
        st.session_state['Status_do_pagamento'] = st.selectbox('Status do Pagamento',lista_status)
        st.session_state['Valor_pago']=st.number_input('Valor pago')
        st.session_state['OBS'] = st.text_input('OBS')

        if st.button('Enviar'):
            cadastrar1(
            st.session_state['Categoria'],
            st.session_state['Conta'],
            st.session_state['Forma_de_pagamento'],
            st.session_state['Data_do_pagamento'],
            st.session_state['Status_do_pagamento'],
            st.session_state['Valor_pago'],
            st.session_state['OBS'])
            st.rerun()
            
            st.success('Cadastro enviado com sucesso!')
        if st.button('Limpar'):
            st.session_state['Categoria']=''
            st.session_state['Conta']=''
            st.session_state['Forma_de_pagamento']=''
            st.session_state['Data_do_pagamento']=''
            st.session_state['Status_do_pagamento']=''
            st.session_state['Valor_pago']=''
            st.session_state['OBS']
            st.rerun()

    if 'Novo gasto' not in st.session_state:
        if st.button('➕'):
            formulario_c_gastos()
    else:
        st.success('Novo gasto enviado com sucesso')

def cadastrar1(Categoria, Conta, Forma_de_pagamento,  Data_do_pagamento, Status_do_pagamento, Valor_pago, OBS):

    nova_linha = {'Categoria': Categoria,
                  'Conta': Conta,
                  'Forma_de_pagamento': Forma_de_pagamento,
                  'Data_do_pagamento': Data_do_pagamento,
                  'Status_do_pagamento': Status_do_pagamento,
                  'Valor_pago': Valor_pago,
                  'OBS': OBS}
    df_nova_linha = pd.DataFrame([nova_linha])
    try:
        wb = load_workbook (filename="Contas.xlsx")
    except FileNotFoundError:
        df_nova_linha.to_excel("Contas.xlsx", index=False)
        wb = load_workbook(filename="Contas.xlsx")
    ws = wb.active
    proxima_linha = ws.max_row + 1

    for index, row in df_nova_linha.iterrows():
        for col, value in enumerate(row, start=1):
            ws.cell(row=proxima_linha, column=col, value=value)
    proxima_linha += 1

    wb.save(filename="Contas.xlsx")

#Criando o Cadastro de nova categoria
def nova_categoria():

    @st.dialog("Cadastrar nova Categoria", width="big")  
    def cadastrar_categoria():
        
        if 'Nova_Categoria' not in st.session_state:
            st.session_state['Nova_Categoria'] = ''
        
        st.session_state['Nova_Categoria'] = st.text_input('Nova categoria', st.session_state['Nova_Categoria'])
        
        if st.button("Enviar"):
            cadastrar2(st.session_state['Nova_Categoria'])
            st.success('Cadastro enviado com sucesso!')
            st.rerun()  
        
        if st.button("Limpar"):
            st.session_state['Nova_Categoria'] = ''
            st.rerun()  
   
    if 'nova_categoria' not in st.session_state:
        if st.button("Abrir Formulário de Categoria"):
            cadastrar_categoria()  
    else:
        st.success("Categoria cadastrada com sucesso!")


def cadastrar2(Nova_Categoria):

    nova_linha2 = {'Nova_Categoria': Nova_Categoria}
    df_nova_linha2 = pd.DataFrame([nova_linha2])
    try:
        wb = load_workbook (filename="nome_categorias.xlsx")
    except FileNotFoundError:
        df_nova_linha2.to_excel("nome_categorias.xlsx", index=False)
        wb = load_workbook(filename="nome_categorias.xlsx")
    ws = wb.active
    proxima_linha = ws.max_row + 1

    for index, row in df_nova_linha2.iterrows():
        for col, value in enumerate(row, start=1):
            ws.cell(row=proxima_linha, column=col, value=value)
    proxima_linha += 1

    wb.save(filename="nome_categorias.xlsx")
    return df


#Criando o Cadastro de nova conta
def Nova_conta():
    
    if 'form_aberto' not in st.session_state:
        st.session_state['form_aberto'] = False

    @st.dialog("Cadastrar nova conta", width="big")  
    def cadastrar_conta():
        
        if 'Nova_conta' not in st.session_state:
            st.session_state['Nova_conta'] = ''
        
        st.session_state['Nova_conta'] = st.text_input('Nova Conta', st.session_state['Nova_conta'])
        
        if st.button("Enviar"):
            cadastrar3(st.session_state['Nova_conta'])
            st.success('Cadastro enviado com sucesso!')
            st.session_state['form_aberto'] = False  
            st.rerun()  

        if st.button("Limpar"):
            st.session_state['Nova_conta'] = ''
            st.session_state['form_aberto'] = False  
            st.rerun()  

    if not st.session_state['form_aberto']:
        if st.button("Abrir Formulário Conta"):
            st.session_state['form_aberto'] = True  
            cadastrar_conta()  


def cadastrar3(Nova_conta):
    nova_linha3 = {'Nova_conta': Nova_conta}
    df_nova_linha3 = pd.DataFrame([nova_linha3])

    try:
        wb = load_workbook(filename="nome_contas.xlsx")
    except FileNotFoundError:
        df_nova_linha3.to_excel("nome_contas.xlsx", index=False)
        wb = load_workbook(filename="nome_contas.xlsx")

    ws = wb.active
    proxima_linha = ws.max_row + 1

    for index, row in df_nova_linha3.iterrows():
        for col, value in enumerate(row, start=1):
            ws.cell(row=proxima_linha, column=col, value=value)

    wb.save(filename="nome_contas.xlsx")


#Criando o Cadastro de recebimentos
def recebimento():
    
    
    if 'form_aberto2' not in st.session_state:
        st.session_state['form_aberto2'] = False

    
    @st.dialog("Cadastrar novo recebimento", width="big")  
    def cadastrar_conta():
        
        lista_recebimento = 'Assim Saúde', 'Colegio QI', 'Pacto Proteção Veicular', 'Outro'

        if 'Fonte_da_recebimento' not in st.session_state:
            st.session_state['Fonte_da_recebimento'] = ''
        if 'Valor_recebido' not in st.session_state:
            st.session_state['Valor_recebido']=''
        if 'Data_Recebimento' not in st.session_state:
            st.session_state['Data_recebimento']=''

        
        st.session_state['Fonte_da_recebimento'] = st.selectbox('Fonte da recebimento',lista_recebimento)
        st.session_state['Valor_recebido'] = st.number_input('Valor Recebido')
        st.session_state['Data_recebimento'] = st.date_input('Data')
        
        if st.button("Enviar"):
            cadastrar4(st.session_state['Fonte_da_recebimento'],
                       st.session_state['Valor_recebido'],
                       st.session_state['Data_recebimento'])
            st.success('Cadastro enviado com sucesso!')
            st.session_state['form_aberto'] = False  
            st.rerun()  

        if st.button("Limpar"):
            st.session_state['Fonte_da_recebimento'] = ''
            st.session_state['form_aberto'] = False  
            st.rerun()  

    if not st.session_state['form_aberto']:
        if st.button("Abrir Formulário nova conta"):
            st.session_state['form_aberto'] = True  
            cadastrar_conta()  


def cadastrar4(Fonte_da_recebimento,Valor_recebido,Data_recebimento):
    nova_linha3 = {'Fonte_da_recebimento': Fonte_da_recebimento,
                   'Valor_recebido': Valor_recebido,
                   'Data_recebimento': Data_recebimento}
    df_nova_linha3 = pd.DataFrame([nova_linha3])

    try:
        wb = load_workbook(filename="Recebimentos.xlsx")
    except FileNotFoundError:
        df_nova_linha3.to_excel("Recebimentos.xlsx", index=False)
        wb = load_workbook(filename="Recebimentos.xlsx")

    ws = wb.active
    proxima_linha = ws.max_row + 1

    for index, row in df_nova_linha3.iterrows():
        for col, value in enumerate(row, start=1):
            ws.cell(row=proxima_linha, column=col, value=value)

    wb.save(filename="Recebimentos.xlsx")









def main():
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False

    if st.session_state['logged_in']:
        pag_inicial()
    else:
        login_page()

if __name__ == "__main__":
    main()    